"""
Fantasy Premier League (FPL) Data Analysis Tool

Fetches, processes, and visualizes FPL data to provide:
- Historical season performance analysis
- Gameweek-by-gameweek statistics
- League participation overview

Outputs:
- Excel report with styled tables and charts
- PNG visualizations of rank progression
"""

import requests
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from datetime import datetime
from matplotlib.ticker import FuncFormatter

from openpyxl.styles import Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import GradientFill, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------
# DATA FETCHING & PROCESSING
# --------------------------

def fetch_fpl_data(team_id=300152):
    """
    Retrieve FPL data from the official API.
    
    Args:
        team_id (int): FPL team identifier (default: 300152)
    
    Returns:
        dict: Dictionary containing:
            - 'team_data': Current team information
            - 'history_data': Historical performance data
    
    Raises:
        Exception: If either API request fails
    """
    base_url = f"https://fantasy.premierleague.com/api/entry/{team_id}/"
    history_url = f"{base_url}history/"
    
    response = requests.get(base_url)
    history_response = requests.get(history_url)
    
    if all([response.ok, history_response.ok]):
        return {
            'team_data': response.json(),
            'history_data': history_response.json()
        }
    raise Exception(f"API Error: Team {response.status_code}, History {history_response.status_code}")

def process_season_data(team_data, history_data):
    """
    Process raw API data into structured seasonal performance DataFrame.
    
    Args:
        team_data (dict): Current team data from API
        history_data (dict): Historical data from API
    
    Returns:
        pd.DataFrame: Combined current and past seasons data with columns:
            - Season (YYYY/YY format)
            - Total Points
            - Rank
            - Bank Balance (£m)
            - Team Value (£m)
            - Status (Current/Past)
    """
    # Current season data (values converted from tenths of £ to £m)
    current_season = {
        'Season': f"{datetime.now().year - 1}/{datetime.now().year % 100:02}",
        'Total Points': team_data['summary_overall_points'],
        'Rank': team_data['summary_overall_rank'],
        'Bank Balance (£m)': team_data['last_deadline_bank'] / 10,
        'Team Value (£m)': history_data['current'][-1]['value'] / 10,
        'Status': 'Current'
    }
    
    # Historical seasons (bank/value not available in API)
    past_seasons = [{
        'Season': s['season_name'],
        'Total Points': s['total_points'],
        'Rank': s['rank'],
        'Bank Balance (£m)': None,
        'Team Value (£m)': None,
        'Status': 'Past'
    } for s in history_data['past']]
    
    # Combine and sort seasons
    season_df = pd.DataFrame([current_season] + past_seasons)
    season_df['SortKey'] = season_df['Season'].str.split('/').str[0].astype(int)
    return season_df.sort_values(by='SortKey', ascending=False).drop('SortKey', axis=1)

def get_current_season_gw_data(history_data):
    """
    Process gameweek data for current season.
    
    Args:
        history_data (dict): Raw API history data
    
    Returns:
        pd.DataFrame: Processed gameweek data with columns:
            - Gameweek
            - Points
            - Total Points
            - Rank
            - Overall Rank
            - Bank Balance (£m)
            - Team Value (£m)
            - Transfers
            - Transfers Cost
            - Points on Bench
    """
    gw_data = pd.DataFrame(history_data['current']).rename(columns={
        'event': 'Gameweek', 
        'points': 'Points', 
        'total_points': 'Total Points',
        'rank': 'Rank', 
        'overall_rank': 'Overall Rank',
        'rank_sort': 'Rank Sort',
        'percentile_rank': 'Percentile Rank',
        'bank': 'Bank Balance (£m)', 
        'value': 'Team Value (£m)',
        'event_transfers': 'Transfers', 
        'event_transfers_cost': 'Transfers Cost', 
        'points_on_bench': 'Points on Bench'
    })
    # Convert from tenths of £ to £m
    gw_data[['Bank Balance (£m)', 'Team Value (£m)']] /= 10
    return gw_data

def process_leagues_data(team_data):
    """
    Extract league participation information.
    
    Args:
        team_data (dict): Raw API team data
    
    Returns:
        pd.DataFrame: League information with columns:
            - League Name
            - Type (Classic/H2H)
            - Your Rank
            - Total Players
            - League ID
            - Admin ID
    """
    leagues = team_data.get('leagues', {})
    return pd.DataFrame([{
        'League Name': l.get('name', 'N/A'),
        'Type': 'Classic' if l in leagues.get('classic', []) else 'H2H',
        'Your Rank': l.get('entry_rank', 'N/A'),
        'Total Players': l.get('total_players', 'N/A'),
        'League ID': l.get('id', 'N/A'),
        'Admin ID': l.get('admin_entry') or "N/A"
    } for l in leagues.get('classic', []) + leagues.get('h2h', [])])

# ---------------------
# EXCEL STYLING & SETUP
# ---------------------

def style_summary_sheet(worksheet, df, best_season):
    """
    Apply professional styling to season summary worksheet.
    
    Args:
        worksheet (Worksheet): Target Excel worksheet
        df (pd.DataFrame): Season data to style
        best_season (pd.Series): Best season row to highlight
    """
    # Header styling configuration
    header_fill = GradientFill(degree=90, stop=("2A5C8A", "1E4A7A"))
    number_formats = {
        'Total Points': '#,##0',
        'Rank': '#,##0',
        'Bank Balance (£m)': '#,##0.0',
        'Team Value (£m)': '#,##0.0'
    }
    
    # Apply header styles and number formatting
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        if col_name in number_formats:
            for row_idx in range(2, len(df)+2):
                worksheet.cell(row=row_idx, column=col_idx).number_format = number_formats[col_name]
    
    # Highlight best season row
    best_row = df[df['Season'] == best_season['Season']].index[0] + 2
    for col in range(1, len(df.columns)+1):
        worksheet.cell(best_row, col).fill = PatternFill(start_color="C6EFCE", fill_type="solid")

def style_gw_sheet(worksheet, df):
    """
    Style gameweek data worksheet with alternating rows and highlights.
    
    Args:
        worksheet (Worksheet): Target Excel worksheet
        df (pd.DataFrame): Gameweek data to style
    """
    header_fill = GradientFill(degree=90, stop=("2A5C8A", "1E4A7A"))
    number_formats = {
        'Points': '#,##0', 
        'Total Points': '#,##0',
        'Rank': '#,##0',
        'Overall Rank': '#,##0',
        'Bank Balance (£m)': '#,##0.0',
        'Team Value (£m)': '#,##0.0',
        'Transfers': '0',
        'Transfers Cost': '0'
    }
    
    vertical_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin')
    )
    
    # Apply vertical borders to all cells
    for row_idx in range(1, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            # Apply vertical borders to all columns
            cell.border = vertical_border
            
            # Header styling
            if row_idx == 1:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                # Remove header bottom border
                cell.border = Border(left=Side(style='thin', color='D8D8D8'),
                                   right=Side(style='thin', color = 'D8D8D8'))
            
            # Number formatting    
            col_name = df.columns[col_idx-1]
            if row_idx > 1 and col_name in number_formats:
                cell.number_format = number_formats[col_name]
                
    # Create alternating row colors
    grey_fill = PatternFill(start_color="E0E0E0", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    
    for row_idx in range(2, len(df) + 2):
        fill = grey_fill if row_idx % 2 == 0 else white_fill
        for col in range(1, len(df.columns) + 1):
            worksheet.cell(row=row_idx, column=col).fill = fill
    
    # Highlight maximum points row
    if not df.empty and 'Points' in df.columns:
        max_points_idx = df['Points'].idxmax()
        if pd.notnull(max_points_idx):
            max_points_row = max_points_idx + 2
            highlight_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
            for col in range(1, len(df.columns)+1):
                worksheet.cell(row=max_points_row, column=col).fill = highlight_fill

def style_leagues_table(worksheet, start_row, df):
    """
    Style leagues table with Excel table features and custom formatting.
    
    Args:
        worksheet (Worksheet): Target Excel worksheet
        start_row (int): Starting row for the table
        df (pd.DataFrame): Leagues data to style
    """
    # Create Excel table structure
    table_start = f"A{start_row + 1}"
    table_end = f"{get_column_letter(len(df.columns))}{start_row + len(df) + 1}"
    
    tab = Table(displayName=f"LeaguesTable_{start_row}", ref=f"{table_start}:{table_end}")
    tab.showFilterButton = True
    
    style = TableStyleInfo(name="TableStyleMedium9",
                          showFirstColumn=False,
                          showLastColumn=False,
                          showRowStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    # Custom header styling
    header_fill = GradientFill(degree=90, stop=("800000", "600000"))
    for col in range(1, len(df.columns)+1):
        cell = worksheet.cell(row=start_row+1, column=col)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    # Apply number formatting
    number_formats = {'Your Rank': '#,##0', 'Total Players': '#,##0',
                     'League ID': '0', 'Admin ID': '0'}
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in number_formats:
            for row_idx in range(start_row+2, start_row+len(df)+2):
                worksheet.cell(row=row_idx, column=col_idx).number_format = number_formats[col_name]

    # Auto-size columns
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(df[col_name].astype(str).apply(len).max(), len(col_name)) + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len
                
def create_rank_progression_chart(worksheet, gw_data):
    """
    Create rank progression line chart in Excel worksheet.
    
    Args:
        worksheet (Worksheet): Target Excel worksheet
        gw_data (pd.DataFrame): Gameweek data containing rank information
    """
    if gw_data.empty:
        return
    
    # Chart configuration
    chart = LineChart()
    chart.title = "Rank Progression"
    chart.style = 6
    chart.height = 10
    chart.width = 27
    chart.legend = None
    chart.y_axis.title = "Overall Rank"
    chart.x_axis.title = "Gameweek"
    chart.y_axis.scaling.orientation = "maxMin"
    chart.y_axis.reverseOrder = True
    chart.y_axis.number_format = '[>=1000000]0.0,,"M";[>=1000]0,"k";0'

    # Data references
    data = Reference(worksheet, min_col=6, min_row=2, max_col=6, max_row=len(gw_data)+1)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(gw_data)+1)

    # Add data and style series
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    series = chart.series[0]
    series.graphicalProperties.line.solidFill = "2A5C8A"
    series.graphicalProperties.line.width = 25000

    # Position chart below GW data
    chart_row = len(gw_data) + 5
    worksheet.add_chart(chart, f"B{chart_row}")

def save_to_excel(main_df, best_season, gw_data, leagues_df, team_id):
    """
    Generate Excel report with formatted worksheets.
    
    Args:
        main_df (pd.DataFrame): Season summary data
        best_season (pd.Series): Best season information
        gw_data (pd.DataFrame): Gameweek data
        leagues_df (pd.DataFrame): Leagues data
        team_id (int): FPL team identifier
    
    Returns:
        str: Output filename
    """
    filename = "myFPL_analysis.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Season Summary Sheet
        main_df.to_excel(writer, sheet_name='Season Summary', index=False)
        ws_summary = writer.sheets['Season Summary']
        style_summary_sheet(ws_summary, main_df, best_season)

        # Gameweek Sheet
        gw_data.to_excel(writer, sheet_name='Current Season GWs', index=False)
        ws_gw = writer.sheets['Current Season GWs']
        style_gw_sheet(ws_gw, gw_data)
        create_rank_progression_chart(ws_gw, gw_data)
    
        # Leagues Table Positioning
        leagues_start_row = len(gw_data) + 3 + 25
        leagues_df.to_excel(writer, 
                          sheet_name='Current Season GWs', 
                          startrow=leagues_start_row, 
                          index=False)
        style_leagues_table(ws_gw, leagues_start_row, leagues_df)
        
        # Column width adjustments
        for column in ws_gw.columns:
            letter = get_column_letter(column[0].column)
            ws_gw.column_dimensions[letter].width = (max(len(str(cell.value)) for cell in column) + 2) * 1.2
        ws_gw.column_dimensions['A'].width = 18
    
    return filename

# ----------------------
# VISUALIZATION & OUTPUT
# ----------------------

def format_rank(x):
    """
    Format rank numbers for human readability.
    
    Args:
        x (int): Raw rank number
    
    Returns:
        str: Formatted rank string (k/M suffixes)
    """
    if x >= 1_000_000:
        return f'{x/1_000_000:.1f}M'
    if x >= 1000:
        return f'{x/1000:.0f}k'
    return f'{x}'

def plot_rank_progression(df, save_path="current_rank_progression.png"):
    """
    Generate and save rank progression line chart.
    
    Args:
        df (pd.DataFrame): Gameweek data with 'Gameweek' and 'Overall Rank'
        save_path (str): Output path for PNG file
    """
    plt.figure(figsize=(24, 6))
    
    # Create base plot
    ax = plt.gca()
    df.plot(x='Gameweek', y='Overall Rank', 
           marker='o', markersize=8,
           linestyle='-', linewidth=2.5,
           color='#2A5C8A', 
           markeredgecolor='white',
           ax=ax)
    
    # Configure titles and labels
    plt.title('Overall Rank Progression', fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Gameweek', fontsize=12, labelpad=12)
    plt.ylabel('Overall Rank', fontsize=12, labelpad=12)
    
    # Remove legend if present
    if ax.get_legend():
        ax.get_legend().remove()
    
    # Axis configuration
    ax.set_xticks(df['Gameweek'])
    ax.invert_yaxis()
    ax.grid(True, linestyle='--', alpha=0.6)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: format_rank(x)))
    
    # Annotate data points
    y_range = df['Overall Rank'].max() - df['Overall Rank'].min()
    label_offset = y_range * 0.08
    prev_rank = None
    
    for idx, row in df.iterrows():
        x = row['Gameweek']
        y = row['Overall Rank']
        
        offset = (-label_offset if (prev_rank and y > prev_rank) 
                 else label_offset) if prev_rank is not None else label_offset
        va = 'top' if (prev_rank and y > prev_rank) else 'bottom'
        
        ax.text(x, y + offset, format_rank(y),
                ha='center', va=va, fontsize=10,
                color='#2c3e50', fontweight='bold',
                bbox=dict(facecolor='white', alpha=0.9,
                        edgecolor='#dddddd', boxstyle='round,pad=0.3'))
        prev_rank = y
    
    # Border styling
    for spine in ax.spines.values():
        spine.set_color('#dddddd')
        spine.set_linewidth(0.5)
    
    # Save output
    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.12)
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()

def plot_season_ranks(df, save_path="season_rank_progression.png"):
    """
    Generate and save historical season ranks line chart.
    
    Args:
        df (pd.DataFrame): Season data with 'Season' and 'Rank'
        save_path (str): Output path for PNG file
    """
    plt.figure(figsize=(24, 6))
    
    # Prepare chronological data
    plot_df = df.copy()
    plot_df['Season Year'] = plot_df['Season'].str.split('/').str[0].astype(int)
    plot_df = plot_df.sort_values('Season Year')
    
    # Create base plot
    ax = plt.gca()
    plot_df.plot(x='Season', y='Rank', 
                marker='o', markersize=8,
                linestyle='-', linewidth=2.5,
                color='#2A5C8A',
                markeredgecolor='white',
                ax=ax)
    
    # Configure titles and labels
    plt.title('Historical Season Rankings', fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Season', fontsize=12, labelpad=12)
    plt.ylabel('Overall Rank', fontsize=12, labelpad=12)
    
    # Remove legend
    if ax.get_legend():
        ax.get_legend().remove()
    
    # Axis configuration
    y_min = (plot_df['Rank'].min() // 200_000) * 200_000
    y_max = ((plot_df['Rank'].max() // 200_000) + 1) * 200_000
    y_ticks = np.arange(y_min, y_max + 200_000, 200_000)
    
    ax.set_yticks(y_ticks)
    ax.invert_yaxis()
    ax.grid(True, linestyle='--', alpha=0.6)
    plt.xticks(rotation=45, ha='right')
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: format_rank(x)))
    
    # Annotate data points
    y_range = plot_df['Rank'].max() - plot_df['Rank'].min()
    label_offset = y_range * 0.08
    prev_rank = None
    
    for idx, row in plot_df.iterrows():
        x_pos = list(plot_df['Season']).index(row['Season'])
        y = row['Rank']
        
        offset = (-label_offset if (prev_rank and y > prev_rank) 
                 else label_offset) if prev_rank is not None else label_offset
        va = 'top' if (prev_rank and y > prev_rank) else 'bottom'
        
        ax.text(x_pos, y + offset, format_rank(y),
                ha='center', va=va, fontsize=10,
                color='#2c3e50', fontweight='bold',
                bbox=dict(facecolor='white', alpha=0.9,
                        edgecolor='#dddddd', boxstyle='round,pad=0.3'))
        prev_rank = y
    
    # Highlight best season
    best_season = plot_df.loc[plot_df['Rank'].idxmin()]
    best_x = list(plot_df['Season']).index(best_season['Season'])
    ax.plot(best_x, best_season['Rank'], 'o',
            markersize=12, markeredgecolor='#2c3e50',
            markerfacecolor='#C6EFCE', markeredgewidth=1.5)
    
    # Border styling
    for spine in ax.spines.values():
        spine.set_color('#dddddd')
        spine.set_linewidth(0.5)
    
    # Save output
    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.25)
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    
# ---------
# EXECUTION
# ---------

def main():
    """Main execution flow for the FPL analysis tool."""
    try:
        team_id = 300152  # TODO: Make configurable via CLI arguments
        print("⚽ Fetching FPL data...")
        data = fetch_fpl_data(team_id)
        
        print("📊 Processing data...")
        season_df = process_season_data(data['team_data'], data['history_data'])
        best_season = season_df.loc[season_df['Rank'].idxmin()]
        gw_data = get_current_season_gw_data(data['history_data'])
        leagues_df = process_leagues_data(data['team_data'])
        
        print("📈 Generating visualizations...")
        plot_rank_progression(gw_data)
        plot_season_ranks(season_df)
        
        print("💾 Generating report...")
        excel_filename = save_to_excel(season_df, best_season, gw_data, leagues_df, team_id)
        
        print(f"\n✅ Successfully created:\n- {excel_filename}\n- current_rank_progression.png\n- season_rank_progression.png")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")

if __name__ == "__main__":
    print("🎮 FPL Advanced Analysis Tool")
    print("=============================\n")
    main()
